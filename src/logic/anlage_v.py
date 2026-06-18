"""Anlage V (Einkünfte aus Vermietung und Verpachtung) – Auswertung & Export.

Aggregiert je Haus und Jahr Mieteinnahmen, Umlagen und sämtliche
Werbungskosten und ergänzt sie um AfA + Erhaltungsaufwand aus den
Investitionen. Liefert eine strukturierte Auswertung sowie einen
Excel- und PDF-Bericht für die Einkommensteuererklärung bzw. den
Steuerberater.
"""

from __future__ import annotations

import html
import sqlite3
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from src.db import buchungen, investitionen, stammdaten
from src.utils.eingaben import betrag_formatieren
from src.utils.paths import exporte_verzeichnis

_WAEHRUNG = '#,##0.00 "€"'

# Einnahme-Kategorien, die als Mieteinnahmen gewertet werden.
_KATEGORIE_MIETE = {"Kaltmiete"}
# Einnahme-Kategorien, die als umlagepflichtige Vorauszahlungen gelten.
_KATEGORIE_UMLAGE = {"Nebenkosten", "Rücklage"}


def _summen_pro_kategorie(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int, typ: str,
) -> list[tuple[str, Decimal]]:
    """Liefert die Jahres-Summen aller Kategorien dieses Typs (sortiert)."""
    zeilen = verbindung.execute(
        "SELECT k.name AS name, b.betrag AS betrag "
        "FROM buchungen b JOIN kategorien k ON k.id = b.kategorie_id "
        "WHERE b.objekt_id = ? AND substr(b.datum, 1, 4) = ? AND k.typ = ?",
        (objekt_id, f"{jahr:04d}", typ),
    ).fetchall()
    summen: dict[str, Decimal] = {}
    for zeile in zeilen:
        try:
            summen[zeile["name"]] = (
                summen.get(zeile["name"], Decimal("0"))
                + Decimal(zeile["betrag"])
            )
        except (ValueError, TypeError):
            continue
    return sorted(summen.items())


def anlage_v_erstellen(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int
) -> dict:
    """Berechnet die Anlage-V-Daten eines Hauses für ein Jahr."""
    haus = verbindung.execute(
        "SELECT name FROM objekte WHERE id = ?", (objekt_id,)
    ).fetchone()
    if haus is None:
        raise ValueError("Haus nicht gefunden.")

    einnahmen = _summen_pro_kategorie(verbindung, objekt_id, jahr, "einnahme")
    ausgaben = _summen_pro_kategorie(verbindung, objekt_id, jahr, "ausgabe")
    afa_info = investitionen.afa_im_jahr(verbindung, objekt_id, jahr)

    miete = sum((b for n, b in einnahmen if n in _KATEGORIE_MIETE),
                Decimal("0"))
    umlagen = sum((b for n, b in einnahmen if n in _KATEGORIE_UMLAGE),
                  Decimal("0"))
    sonstige_einnahmen = sum(
        (b for n, b in einnahmen
         if n not in _KATEGORIE_MIETE and n not in _KATEGORIE_UMLAGE),
        Decimal("0")
    )
    laufende_ausgaben = sum((b for _, b in ausgaben), Decimal("0"))
    erhaltung = afa_info["erhaltung"]
    afa = afa_info["afa"]

    summe_einnahmen = miete + umlagen + sonstige_einnahmen
    summe_werbungskosten = laufende_ausgaben + erhaltung + afa
    einkuenfte = summe_einnahmen - summe_werbungskosten

    return {
        "haus": haus["name"],
        "jahr": jahr,
        "mieteinnahmen": miete,
        "umlagen": umlagen,
        "sonstige_einnahmen": sonstige_einnahmen,
        "summe_einnahmen": summe_einnahmen,
        "laufende_ausgaben": ausgaben,
        "laufende_ausgaben_summe": laufende_ausgaben,
        "erhaltung": erhaltung,
        "afa": afa,
        "investitionspositionen": afa_info["positionen"],
        "summe_werbungskosten": summe_werbungskosten,
        "einkuenfte": einkuenfte,
    }


def anlage_v_excel(verbindung: sqlite3.Connection, jahr: int) -> Path:
    """Schreibt eine Anlage-V-Übersicht für alle Häuser des Jahres als Excel."""
    wb = Workbook()
    blatt = wb.active
    blatt.title = f"Anlage V {jahr}"
    fett = Font(bold=True)
    blatt.column_dimensions["A"].width = 36
    for sp in ("B", "C", "D", "E"):
        blatt.column_dimensions[sp].width = 16

    zeile = 1
    titel = blatt.cell(zeile, 1, f"Anlage V {jahr} — Einkünfte aus V+V")
    titel.font = Font(bold=True, size=14)
    zeile += 2

    gesamt_einkuenfte = Decimal("0")
    for haus in stammdaten.objekte_laden(verbindung):
        daten = anlage_v_erstellen(verbindung, haus["id"], jahr)

        kopf = blatt.cell(zeile, 1, daten["haus"])
        kopf.font = Font(bold=True, size=12)
        zeile += 1

        blatt.cell(zeile, 1, "Einnahmen").font = fett
        zeile += 1
        for label, wert in (
            ("Mieteinnahmen (Kaltmiete)", daten["mieteinnahmen"]),
            ("Umlagen (Nebenkosten, Rücklage)", daten["umlagen"]),
            ("Sonstige Einnahmen", daten["sonstige_einnahmen"]),
        ):
            blatt.cell(zeile, 1, label)
            zelle = blatt.cell(zeile, 2, float(wert))
            zelle.number_format = _WAEHRUNG
            zeile += 1
        blatt.cell(zeile, 1, "Summe Einnahmen").font = fett
        s = blatt.cell(zeile, 2, float(daten["summe_einnahmen"]))
        s.number_format = _WAEHRUNG
        s.font = fett
        zeile += 2

        blatt.cell(zeile, 1, "Werbungskosten").font = fett
        zeile += 1
        for name, betrag in daten["laufende_ausgaben"]:
            blatt.cell(zeile, 1, name)
            zelle = blatt.cell(zeile, 2, float(betrag))
            zelle.number_format = _WAEHRUNG
            zeile += 1
        if daten["erhaltung"] > 0:
            blatt.cell(zeile, 1, "Erhaltungsaufwand (Reparaturen)")
            zelle = blatt.cell(zeile, 2, float(daten["erhaltung"]))
            zelle.number_format = _WAEHRUNG
            zeile += 1
        if daten["afa"] > 0:
            blatt.cell(zeile, 1, "AfA (Abschreibung Herstellungsaufwand)")
            zelle = blatt.cell(zeile, 2, float(daten["afa"]))
            zelle.number_format = _WAEHRUNG
            zeile += 1
        blatt.cell(zeile, 1, "Summe Werbungskosten").font = fett
        s = blatt.cell(zeile, 2, float(daten["summe_werbungskosten"]))
        s.number_format = _WAEHRUNG
        s.font = fett
        zeile += 2

        blatt.cell(zeile, 1, "Einkünfte aus V+V").font = fett
        s = blatt.cell(zeile, 2, float(daten["einkuenfte"]))
        s.number_format = _WAEHRUNG
        s.font = fett
        zeile += 3

        gesamt_einkuenfte += daten["einkuenfte"]

    blatt.cell(zeile, 1, "Gesamt-Einkünfte aus V+V").font = Font(
        bold=True, size=13
    )
    s = blatt.cell(zeile, 2, float(gesamt_einkuenfte))
    s.number_format = _WAEHRUNG
    s.font = Font(bold=True, size=13)

    ziel = exporte_verzeichnis() / f"AnlageV_{jahr}.xlsx"
    ziel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ziel)
    return ziel


def steuerbericht_html(daten: dict) -> str:
    """Erzeugt den HTML-Inhalt für den Steuerberater-PDF-Bericht."""

    def zeile_kosten(name: str, wert: Decimal) -> str:
        return (
            f"<tr><td>{html.escape(name)}</td>"
            f"<td align='right'>{betrag_formatieren(wert)} €</td></tr>"
        )

    laufend = "".join(
        zeile_kosten(name, betrag)
        for name, betrag in daten["laufende_ausgaben"]
    )

    investitionen_html = ""
    if daten["investitionspositionen"]:
        rows = "".join(
            f"<tr><td>{html.escape(p['datum'])}</td>"
            f"<td>{html.escape(p['beschreibung'])}</td>"
            f"<td>{html.escape(p['typ'])}</td>"
            f"<td align='right'>{betrag_formatieren(p['anteil'])} €</td>"
            f"<td align='right'>{betrag_formatieren(p['gesamt'])} €</td></tr>"
            for p in daten["investitionspositionen"]
        )
        investitionen_html = (
            "<h3>Investitionen / AfA</h3>"
            "<table border='1' cellspacing='0' cellpadding='4' width='100%'>"
            "<tr><th>Datum</th><th>Beschreibung</th><th>Typ</th>"
            "<th align='right'>Jahres-Anteil</th>"
            "<th align='right'>Gesamtbetrag</th></tr>"
            + rows + "</table>"
        )

    return f"""
    <h2>Anlage V {daten['jahr']} — {html.escape(daten['haus'])}</h2>
    <h3>Einnahmen</h3>
    <table border='1' cellspacing='0' cellpadding='4' width='100%'>
      <tr><td>Mieteinnahmen (Kaltmiete)</td>
        <td align='right'>{betrag_formatieren(daten['mieteinnahmen'])} €</td></tr>
      <tr><td>Umlagen (Nebenkosten, Rücklage)</td>
        <td align='right'>{betrag_formatieren(daten['umlagen'])} €</td></tr>
      <tr><td>Sonstige Einnahmen</td>
        <td align='right'>{betrag_formatieren(daten['sonstige_einnahmen'])} €</td></tr>
      <tr><td><b>Summe Einnahmen</b></td>
        <td align='right'><b>{betrag_formatieren(daten['summe_einnahmen'])} €</b></td></tr>
    </table>

    <h3>Werbungskosten</h3>
    <table border='1' cellspacing='0' cellpadding='4' width='100%'>
      {laufend}
      <tr><td>Erhaltungsaufwand</td>
        <td align='right'>{betrag_formatieren(daten['erhaltung'])} €</td></tr>
      <tr><td>AfA (Abschreibung)</td>
        <td align='right'>{betrag_formatieren(daten['afa'])} €</td></tr>
      <tr><td><b>Summe Werbungskosten</b></td>
        <td align='right'><b>{betrag_formatieren(daten['summe_werbungskosten'])} €</b></td></tr>
    </table>

    <h3>Einkünfte aus Vermietung und Verpachtung</h3>
    <p style='font-size: 14pt;'>
      <b>{betrag_formatieren(daten['einkuenfte'])} €</b>
    </p>

    {investitionen_html}

    <p><i>Hinweis: Diese Übersicht wurde automatisch aus der
    Hausverwaltungs-Software erzeugt und ist vor Einreichung beim
    Finanzamt zu prüfen.</i></p>
    """


def steuerbericht_pdfs(
    verbindung: sqlite3.Connection, jahr: int
) -> list[Path]:
    """Erstellt je Haus einen PDF-Steuerbericht für das Jahr."""
    from src.logic.nebenkosten import _dateiname_saeubern, html_nach_pdf

    ordner = exporte_verzeichnis()
    ordner.mkdir(parents=True, exist_ok=True)
    erstellt: list[Path] = []
    for haus in stammdaten.objekte_laden(verbindung):
        daten = anlage_v_erstellen(verbindung, haus["id"], jahr)
        ziel = ordner / (
            f"Steuerbericht_{_dateiname_saeubern(daten['haus'])}_{jahr}.pdf"
        )
        html_nach_pdf(steuerbericht_html(daten), ziel)
        erstellt.append(ziel)
    return erstellt
