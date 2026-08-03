"""Nebenkostenabrechnung: Berechnung und Export (Excel + PDF).

Je Haus werden die als umlagefähig markierten Ausgaben eines Jahres
zusammengezählt und nach dem Umlageschlüssel des Hauses auf die Mieter
verteilt. Aus den geleisteten Vorauszahlungen ergibt sich je Mieter
eine Nachzahlung oder ein Guthaben.

Bei unterjährigem Ein- oder Auszug wird der Umlageschlüssel taggenau
mit dem Nutzungsanteil gewichtet (z. B. 60 m² × 184/365 Tage).

Vereinfachungen (bewusst, für die private Nutzung):
* Abrechnungszeitraum ist stets das Kalenderjahr.
* Heizkosten werden wie alle übrigen umlagefähigen Kosten verteilt
  (keine verbrauchsabhängige Abrechnung nach Heizkostenverordnung).
* Der Anteil ergibt sich aus gewichteter Schlüssel ÷ Summe der
  gewichteten Schlüssel; Leerstand verteilt sich damit auf die
  verbleibenden Mieter.
"""

from __future__ import annotations

import html
import sqlite3
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from src.db.mietzahlungen import bezahlte_monate
from src.db.stammdaten import UMLAGESCHLUESSEL, objekte_laden
from src.utils.eingaben import ValidierungsFehler, betrag_formatieren
from src.utils.paths import exporte_verzeichnis

_CENT = Decimal("0.01")
_WAEHRUNG = '#,##0.00 "€"'


def _datum_lesen(text: str | None) -> date | None:
    """Liest ein ISO-Datum (YYYY-MM-DD) tolerant ein."""
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def zeitanteil_im_jahr(
    aktiv_von: str | None, aktiv_bis: str | None, jahr: int
) -> tuple[Decimal, date | None, date | None]:
    """Berechnet den taggenauen Nutzungsanteil eines Mieters am Jahr.

    Liefert ``(anteil, nutzung_von, nutzung_bis)``: der Anteil liegt
    zwischen 0 (Zeitraum berührt das Jahr nicht) und 1 (ganzjährig);
    die beiden Daten begrenzen den tatsächlichen Nutzungszeitraum
    innerhalb des Jahres. Bei unterjährigem Ein-/Auszug trägt der
    Mieter nur den Anteil seiner Nutzungstage.
    """
    jahr_start = date(jahr, 1, 1)
    jahr_ende = date(jahr, 12, 31)
    von = _datum_lesen(aktiv_von) or jahr_start
    bis = _datum_lesen(aktiv_bis) or jahr_ende
    start = max(von, jahr_start)
    ende = min(bis, jahr_ende)
    if ende < start:
        return Decimal("0"), None, None
    tage = (ende - start).days + 1
    jahr_tage = (jahr_ende - jahr_start).days + 1
    return Decimal(tage) / Decimal(jahr_tage), start, ende


def umlagefaehige_buchungen(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int
) -> list[dict]:
    """Liefert die einzelnen umlagefähigen Buchungen eines Hauses im Jahr.

    Jeder Eintrag enthält Datum, Kategorie, Beschreibung, Betrag und ob
    ein Beleg hinterlegt ist — die Grundlage für die Belegliste der
    Abrechnung und für den Hinweis, wo noch Belege fehlen.
    """
    zeilen = verbindung.execute(
        "SELECT b.id, b.datum, b.betrag, b.beschreibung, b.beleg_pfad, "
        "k.name AS kategorie_name "
        "FROM buchungen b JOIN kategorien k ON k.id = b.kategorie_id "
        "WHERE b.objekt_id = ? AND substr(b.datum, 1, 4) = ? "
        "AND k.typ = 'ausgabe' AND k.umlagefaehig = 1 "
        "ORDER BY b.datum, b.id",
        (objekt_id, f"{jahr:04d}"),
    ).fetchall()
    ergebnis: list[dict] = []
    for zeile in zeilen:
        try:
            betrag = Decimal(zeile["betrag"])
        except (ValueError, TypeError):
            continue
        ergebnis.append({
            "id": zeile["id"],
            "datum": zeile["datum"],
            "kategorie": zeile["kategorie_name"],
            "beschreibung": zeile["beschreibung"] or "",
            "betrag": betrag,
            "beleg": bool(zeile["beleg_pfad"]),
        })
    return ergebnis


def _umlagefaehige_kosten(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int
) -> list[tuple[str, Decimal]]:
    """Liefert die umlagefähigen Ausgaben eines Hauses je Kategorie."""
    zeilen = verbindung.execute(
        "SELECT k.name AS name, b.betrag AS betrag "
        "FROM buchungen b JOIN kategorien k ON k.id = b.kategorie_id "
        "WHERE b.objekt_id = ? AND substr(b.datum, 1, 4) = ? "
        "AND k.typ = 'ausgabe' AND k.umlagefaehig = 1",
        (objekt_id, f"{jahr:04d}"),
    ).fetchall()
    summen: dict[str, Decimal] = {}
    for zeile in zeilen:
        try:
            betrag = Decimal(zeile["betrag"])
        except (ValueError, TypeError):
            continue
        summen[zeile["name"]] = summen.get(zeile["name"], Decimal("0")) + betrag
    return sorted(summen.items())


def abrechnung_erstellen(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int
) -> dict:
    """Berechnet die Nebenkostenabrechnung eines Hauses für ein Jahr."""
    haus = verbindung.execute(
        "SELECT name, umlageschluessel FROM objekte WHERE id = ?",
        (objekt_id,),
    ).fetchone()
    if haus is None:
        raise ValidierungsFehler("Das Haus wurde nicht gefunden.")
    schluessel = haus["umlageschluessel"]

    kosten = _umlagefaehige_kosten(verbindung, objekt_id, jahr)
    kosten_gesamt = sum((betrag for _, betrag in kosten), Decimal("0"))

    alle_mieter = verbindung.execute(
        "SELECT id, name, nebenkosten, aktiv_von, aktiv_bis, wohnflaeche, "
        "personenzahl, umlage_gewicht FROM mieter WHERE objekt_id = ? "
        "ORDER BY name COLLATE NOCASE",
        (objekt_id,),
    ).fetchall()

    def schluesselwert(zeile: sqlite3.Row) -> Decimal:
        # Manuell gesetztes Umlage-Gewicht hat Vorrang (Aufteilung
        # bearbeiten): überschreibt den automatischen Schlüssel.
        try:
            manuell = zeile["umlage_gewicht"]
        except (KeyError, IndexError):
            manuell = None
        if manuell is not None and str(manuell).strip() != "":
            try:
                return Decimal(str(manuell))
            except (ValueError, TypeError):
                pass
        if schluessel == "personen":
            return Decimal(zeile["personenzahl"])
        if schluessel == "gleich":
            return Decimal("1")
        try:
            return Decimal(str(zeile["wohnflaeche"]))
        except (ValueError, TypeError):
            return Decimal("0")

    # Taggenauer Nutzungsanteil je Mieter: bei unterjährigem Ein- oder
    # Auszug wird der Umlageschlüssel mit dem Zeitanteil gewichtet
    # (z. B. 60 m² × ein halbes Jahr = 30 gewichtete Anteile).
    mieter: list[tuple[sqlite3.Row, Decimal, date, date]] = []
    for m in alle_mieter:
        anteil_zeit, von, bis = zeitanteil_im_jahr(
            m["aktiv_von"], m["aktiv_bis"], jahr
        )
        if anteil_zeit > 0:
            mieter.append((m, anteil_zeit, von, bis))

    summe_schluessel = sum(
        (schluesselwert(m) * anteil_zeit for m, anteil_zeit, _, _ in mieter),
        Decimal("0"),
    )

    mieter_ergebnis = []
    for m, anteil_zeit, von, bis in mieter:
        wert = schluesselwert(m) * anteil_zeit
        anteil = (wert / summe_schluessel) if summe_schluessel > 0 \
            else Decimal("0")
        kostenanteil = (kosten_gesamt * anteil).quantize(_CENT)
        monate = len(bezahlte_monate(verbindung, m["id"], jahr))
        try:
            monatlich = Decimal(str(m["nebenkosten"]))
        except (ValueError, TypeError):
            monatlich = Decimal("0")
        vorauszahlung = (monatlich * monate).quantize(_CENT)
        mieter_ergebnis.append({
            "name": m["name"],
            "schluesselwert": wert,
            "anteil": anteil,
            "zeitanteil": anteil_zeit,
            "nutzung_von": von.isoformat(),
            "nutzung_bis": bis.isoformat(),
            "kostenanteil": kostenanteil,
            "monate": monate,
            "vorauszahlung": vorauszahlung,
            "differenz": vorauszahlung - kostenanteil,
        })

    einzelbuchungen = umlagefaehige_buchungen(verbindung, objekt_id, jahr)
    return {
        "haus": haus["name"],
        "jahr": jahr,
        "umlageschluessel": schluessel,
        "kosten": kosten,
        "kosten_gesamt": kosten_gesamt,
        "mieter": mieter_ergebnis,
        "einzelbuchungen": einzelbuchungen,
        "belege_fehlen": sum(1 for b in einzelbuchungen if not b["beleg"]),
    }


def abrechnung_excel(verbindung: sqlite3.Connection, jahr: int) -> Path:
    """Erstellt eine Excel-Datei mit der Nebenkostenabrechnung aller Häuser."""
    arbeitsmappe = Workbook()
    blatt = arbeitsmappe.active
    blatt.title = f"Nebenkosten {jahr}"
    fett = Font(bold=True)

    blatt.column_dimensions["A"].width = 28
    for spalte in ("B", "C", "D", "E"):
        blatt.column_dimensions[spalte].width = 16

    zeile = 1
    titel = blatt.cell(zeile, 1, f"Nebenkostenabrechnung {jahr}")
    titel.font = Font(bold=True, size=14)
    zeile += 2

    for haus in objekte_laden(verbindung):
        abrechnung = abrechnung_erstellen(verbindung, haus["id"], jahr)
        kopf = blatt.cell(
            zeile, 1,
            f"{abrechnung['haus']} "
            f"({UMLAGESCHLUESSEL[abrechnung['umlageschluessel']]})",
        )
        kopf.font = Font(bold=True, size=12)
        zeile += 1

        blatt.cell(zeile, 1, "Umlagefähige Kosten").font = fett
        zeile += 1
        for name, betrag in abrechnung["kosten"]:
            blatt.cell(zeile, 1, name)
            zelle = blatt.cell(zeile, 2, float(betrag))
            zelle.number_format = _WAEHRUNG
            zeile += 1
        blatt.cell(zeile, 1, "Summe").font = fett
        summe = blatt.cell(zeile, 2, float(abrechnung["kosten_gesamt"]))
        summe.number_format = _WAEHRUNG
        summe.font = fett
        zeile += 2

        for index, titeltext in enumerate(
            ("Mieter", "Anteil", "Kostenanteil", "Vorauszahlung", "Differenz")
        ):
            blatt.cell(zeile, 1 + index, titeltext).font = fett
        zeile += 1
        for eintrag in abrechnung["mieter"]:
            blatt.cell(zeile, 1, eintrag["name"])
            blatt.cell(zeile, 2, f"{eintrag['anteil'] * 100:.1f} %")
            for spalte, schluessel in (
                (3, "kostenanteil"), (4, "vorauszahlung"), (5, "differenz")
            ):
                zelle = blatt.cell(zeile, spalte, float(eintrag[schluessel]))
                zelle.number_format = _WAEHRUNG
            zeile += 1
        zeile += 2

    ziel = exporte_verzeichnis() / f"Nebenkosten_{jahr}.xlsx"
    ziel.parent.mkdir(parents=True, exist_ok=True)
    arbeitsmappe.save(ziel)
    return ziel


def abrechnung_html(abrechnung: dict, eintrag: dict) -> str:
    """Erzeugt den HTML-Inhalt der Abrechnung für einen einzelnen Mieter."""
    schluessel_text = UMLAGESCHLUESSEL[abrechnung["umlageschluessel"]]
    kosten_zeilen = "".join(
        f"<tr><td>{html.escape(name)}</td>"
        f"<td align='right'>{betrag_formatieren(betrag)} €</td></tr>"
        for name, betrag in abrechnung["kosten"]
    )
    differenz = eintrag["differenz"]
    if differenz >= 0:
        fazit = (f"Guthaben — Rückzahlung an den Mieter: "
                 f"{betrag_formatieren(differenz)} €")
    else:
        fazit = (f"Nachzahlung durch den Mieter: "
                 f"{betrag_formatieren(-differenz)} €")

    def datum_de(iso: str) -> str:
        return f"{iso[8:10]}.{iso[5:7]}.{iso[:4]}"

    # Belegliste: jede einzelne umlagefähige Buchung des Hauses.
    einzel_zeilen = "".join(
        f"<tr><td>{datum_de(b['datum'])}</td>"
        f"<td>{html.escape(b['kategorie'])}</td>"
        f"<td>{html.escape(b['beschreibung'][:70])}</td>"
        f"<td align='right'>{betrag_formatieren(b['betrag'])} €</td>"
        f"<td align='center'>{'✓' if b['beleg'] else '—'}</td></tr>"
        for b in abrechnung.get("einzelbuchungen", [])
    )
    belegliste = ""
    if einzel_zeilen:
        belegliste = (
            "<h3>Kostenaufstellung im Einzelnen</h3>"
            "<table border='1' cellspacing='0' cellpadding='4' width='100%'>"
            "<tr><th>Datum</th><th>Kategorie</th><th>Beschreibung</th>"
            "<th align='right'>Betrag</th><th>Beleg</th></tr>"
            + einzel_zeilen +
            "</table>"
            "<p><i>Die Belege können nach Terminabsprache eingesehen "
            "werden.</i></p>"
        )

    zeitanteil = eintrag.get("zeitanteil", Decimal("1"))
    nutzungszeile = ""
    if zeitanteil < 1:
        nutzungszeile = (
            "<tr><td>Ihr Nutzungszeitraum (taggenau)</td>"
            f"<td align='right'>{datum_de(eintrag['nutzung_von'])} – "
            f"{datum_de(eintrag['nutzung_bis'])} "
            f"({zeitanteil * 100:.1f} % des Jahres)</td></tr>"
        )

    return f"""
    <h2>Nebenkostenabrechnung {abrechnung['jahr']}</h2>
    <p>
      <b>Haus:</b> {html.escape(abrechnung['haus'])}<br>
      <b>Mieter:</b> {html.escape(eintrag['name'])}<br>
      <b>Abrechnungszeitraum:</b>
      01.01.{abrechnung['jahr']} – 31.12.{abrechnung['jahr']}<br>
      <b>Umlageschlüssel:</b> {schluessel_text}
    </p>
    <h3>Umlagefähige Gesamtkosten des Hauses</h3>
    <table border='1' cellspacing='0' cellpadding='5' width='100%'>
      {kosten_zeilen}
      <tr><td><b>Summe</b></td><td align='right'><b>
      {betrag_formatieren(abrechnung['kosten_gesamt'])} €</b></td></tr>
    </table>
    <h3>Ihr Anteil</h3>
    <table border='1' cellspacing='0' cellpadding='5' width='100%'>
      {nutzungszeile}
      <tr><td>Anteil ({schluessel_text})</td>
        <td align='right'>{eintrag['anteil'] * 100:.1f} %</td></tr>
      <tr><td>Ihr Kostenanteil</td>
        <td align='right'>{betrag_formatieren(eintrag['kostenanteil'])} €</td></tr>
      <tr><td>Geleistete Vorauszahlung ({eintrag['monate']} Monate)</td>
        <td align='right'>
        {betrag_formatieren(eintrag['vorauszahlung'])} €</td></tr>
    </table>
    <h3>{fazit}</h3>
    {belegliste}
    <p><i>Hinweis: Diese Abrechnung wurde automatisch erstellt und ist
    vor Weitergabe zu prüfen.</i></p>
    """


def html_nach_pdf(
    inhalt_html: str, ziel_pfad: Path, quer: bool = False
) -> None:
    """Schreibt einen HTML-Inhalt als PDF-Datei (über Qt, ohne Zusatzpaket).

    ``quer=True`` erzeugt A4-Querformat (für breite Tabellen).
    """
    from PySide6.QtGui import (
        QPageLayout, QPageSize, QPdfWriter, QTextDocument,
    )

    schreiber = QPdfWriter(str(ziel_pfad))
    schreiber.setPageSize(QPageSize(QPageSize.A4))
    if quer:
        schreiber.setPageOrientation(QPageLayout.Landscape)
    dokument = QTextDocument()
    dokument.setHtml(inhalt_html)
    dokument.print_(schreiber)


def _dateiname_saeubern(text: str) -> str:
    """Entfernt für Dateinamen ungeeignete Zeichen."""
    erlaubt = [z if z.isalnum() or z in " -_" else "_" for z in text]
    return "".join(erlaubt).strip().replace(" ", "_")


def abrechnung_pdfs(
    verbindung: sqlite3.Connection, objekt_id: int, jahr: int
) -> list[Path]:
    """Erstellt je Mieter eines Hauses eine PDF-Abrechnung."""
    abrechnung = abrechnung_erstellen(verbindung, objekt_id, jahr)
    ordner = exporte_verzeichnis()
    ordner.mkdir(parents=True, exist_ok=True)

    erstellt: list[Path] = []
    haus_kurz = _dateiname_saeubern(abrechnung["haus"])
    for eintrag in abrechnung["mieter"]:
        mieter_kurz = _dateiname_saeubern(eintrag["name"])
        ziel = ordner / f"Nebenkosten_{haus_kurz}_{mieter_kurz}_{jahr}.pdf"
        html_nach_pdf(abrechnung_html(abrechnung, eintrag), ziel)
        erstellt.append(ziel)
    return erstellt


# =========================================================================
# Mieteinnahmen-Nachweis (für die Steuer / Anlage V)
# =========================================================================

_MONATE_KURZ = [
    "Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
    "Jul", "Aug", "Sep", "Okt", "Nov", "Dez",
]


def mieteinnahmen_daten(
    verbindung: sqlite3.Connection, objekt_id: int | None, jahr: int
) -> dict:
    """Sammelt die erfassten Mietzahlungen je Mieter für ein Jahr.

    Grundlage für den Steuer-Nachweis „wer wann wie viel gezahlt hat".
    ``objekt_id=None`` wertet alle Häuser aus. Für jeden Mieter werden
    die Monatsbeträge (aus der Mietzahlungs-Checkliste) und die Summe
    geliefert, dazu Soll-Miete und die Häuser.
    """
    haus_filter = ""
    params: list = [f"{jahr:04d}"]
    if objekt_id is not None:
        haus_filter = "AND m.objekt_id = ?"
        params.append(objekt_id)

    zeilen = verbindung.execute(
        "SELECT m.id AS mieter_id, m.name AS mieter, o.name AS haus, "
        "m.kaltmiete, m.nebenkosten, m.ruecklage, z.monat, z.betrag "
        "FROM mietzahlungen z "
        "JOIN mieter m ON m.id = z.mieter_id "
        "JOIN objekte o ON o.id = m.objekt_id "
        f"WHERE z.jahr = ? {haus_filter} "
        "ORDER BY o.name COLLATE NOCASE, m.name COLLATE NOCASE, z.monat",
        params,
    ).fetchall()

    mieter: dict[int, dict] = {}
    for z in zeilen:
        eintrag = mieter.setdefault(z["mieter_id"], {
            "mieter": z["mieter"],
            "haus": z["haus"],
            "monate": {},
            "summe": Decimal("0"),
            "soll_monat": _summe_soll(z),
        })
        try:
            betrag = Decimal(str(z["betrag"]))
        except (ValueError, TypeError):
            betrag = Decimal("0")
        eintrag["monate"][z["monat"]] = betrag
        eintrag["summe"] += betrag

    liste = sorted(mieter.values(), key=lambda e: (e["haus"], e["mieter"]))
    gesamt = sum((e["summe"] for e in liste), Decimal("0"))
    return {"jahr": jahr, "mieter": liste, "gesamt": gesamt}


def _summe_soll(zeile: sqlite3.Row) -> Decimal:
    """Summiert die vereinbarte Monatsmiete (Kalt + NK + Rücklage)."""
    gesamt = Decimal("0")
    for feld in ("kaltmiete", "nebenkosten", "ruecklage"):
        try:
            gesamt += Decimal(str(zeile[feld]))
        except (ValueError, TypeError):
            pass
    return gesamt


def mieteinnahmen_html(daten: dict, haus_name: str) -> str:
    """Baut den Mieteinnahmen-Nachweis als HTML (Mieter × Monate)."""
    kopf = "".join(f"<th>{m}</th>" for m in _MONATE_KURZ)
    zeilen_html = ""
    for e in daten["mieter"]:
        zellen = ""
        for monat in range(1, 13):
            betrag = e["monate"].get(monat)
            if betrag is not None:
                zellen += (f"<td align='right'>"
                           f"{betrag_formatieren(betrag)}</td>")
            else:
                zellen += "<td align='center'>–</td>"
        zeilen_html += (
            f"<tr><td>{html.escape(e['mieter'])}</td>"
            f"<td>{html.escape(e['haus'])}</td>"
            f"{zellen}"
            f"<td align='right'><b>{betrag_formatieren(e['summe'])}</b></td>"
            f"</tr>"
        )

    return f"""
    <h2>Mieteinnahmen-Nachweis {daten['jahr']}</h2>
    <p><b>Haus:</b> {html.escape(haus_name)}<br>
    <b>Zeitraum:</b> 01.01.{daten['jahr']} – 31.12.{daten['jahr']}</p>
    <p>Beträge in Euro, wie in der Mietzahlungs-Checkliste erfasst
    (– = kein Zahlungseingang vermerkt).</p>
    <table border='1' cellspacing='0' cellpadding='3' width='100%'>
      <tr><th>Mieter</th><th>Haus</th>{kopf}<th>Summe</th></tr>
      {zeilen_html}
      <tr><td colspan='14'></td>
      <td align='right'><b>{betrag_formatieren(daten['gesamt'])}</b></td></tr>
    </table>
    <p><i>Nachweis für die Anlage V (Einkünfte aus Vermietung und
    Verpachtung). Automatisch erstellt, bitte vor Abgabe prüfen.</i></p>
    """


def mieteinnahmen_pdf(
    verbindung: sqlite3.Connection, objekt_id: int | None, jahr: int,
    haus_name: str,
) -> Path:
    """Erstellt den Mieteinnahmen-Nachweis eines Jahres als PDF."""
    daten = mieteinnahmen_daten(verbindung, objekt_id, jahr)
    ordner = exporte_verzeichnis()
    ordner.mkdir(parents=True, exist_ok=True)
    ziel = ordner / (
        f"Mieteinnahmen_{_dateiname_saeubern(haus_name)}_{jahr}.pdf"
    )
    html_nach_pdf(mieteinnahmen_html(daten, haus_name), ziel, quer=True)
    return ziel
