"""Excel-Export der Jahresübersicht mit openpyxl.

Je Haus wird ein Block geschrieben: Einnahmen- und Ausgaben-Kategorien
mit Monatsspalten, gefolgt von einem Mieter-Block mit dem Status der
Mietzahlungen. Summen- und Saldo-Zeilen werden als Excel-Formeln
(``=SUM(...)``) ausgegeben, nicht als feste Werte.
"""

from __future__ import annotations

import sqlite3
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.db.mietzahlungen import bezahlte_monate
from src.db.stammdaten import kategorien_laden, mieter_laden, objekte_laden
from src.utils.eingaben import MONATSNAMEN_KURZ
from src.utils.paths import exporte_verzeichnis

# Zahlenformat für Geldbeträge.
_WAEHRUNG = '#,##0.00 "€"'
# Erste/letzte Monatsspalte (B = 2 … M = 13), Jahresspalte N = 14.
_ERSTE_MONATSSPALTE = 2
_JAHRESSPALTE = 14


def _monatssummen(
    verbindung: sqlite3.Connection, jahr: int
) -> dict[tuple, dict[int, Decimal]]:
    """Aggregiert die Buchungsbeträge je (Haus, Kategorie) und Monat."""
    zeilen = verbindung.execute(
        "SELECT objekt_id, kategorie_id, substr(datum, 6, 2) AS monat, betrag "
        "FROM buchungen WHERE substr(datum, 1, 4) = ?",
        (f"{jahr:04d}",),
    ).fetchall()
    summen: dict[tuple, dict[int, Decimal]] = {}
    for zeile in zeilen:
        try:
            monat = int(zeile["monat"])
            betrag = Decimal(zeile["betrag"])
        except (ValueError, TypeError):
            continue
        if not 1 <= monat <= 12:
            continue
        schluessel = (zeile["objekt_id"], zeile["kategorie_id"])
        eintrag = summen.setdefault(schluessel, {})
        eintrag[monat] = eintrag.get(monat, Decimal("0")) + betrag
    return summen


def _kopfzeile(arbeitsblatt, zeile: int, erste_spalte_text: str) -> None:
    """Schreibt eine Kopfzeile mit Monatsnamen und Jahresspalte."""
    fett = Font(bold=True)
    kopf = arbeitsblatt.cell(zeile, 1, erste_spalte_text)
    kopf.font = fett
    for index, name in enumerate(MONATSNAMEN_KURZ):
        arbeitsblatt.cell(zeile, _ERSTE_MONATSSPALTE + index, name).font = fett
    arbeitsblatt.cell(zeile, _JAHRESSPALTE, "Jahr").font = fett


def _kategorie_block(
    arbeitsblatt,
    zeile: int,
    titel: str,
    kategorien: list,
    haus_id: int,
    summen: dict,
) -> tuple[int, int]:
    """Schreibt einen Kategorieblock und liefert (nächste Zeile, Summenzeile)."""
    fett = Font(bold=True)
    arbeitsblatt.cell(zeile, 1, titel).font = fett
    zeile += 1

    erste_datenzeile = zeile
    for kategorie in kategorien:
        arbeitsblatt.cell(zeile, 1, kategorie["name"])
        monate = summen.get((haus_id, kategorie["id"]), {})
        for monat in range(1, 13):
            wert = monate.get(monat, Decimal("0"))
            zelle = arbeitsblatt.cell(
                zeile, _ERSTE_MONATSSPALTE + monat - 1, float(wert)
            )
            zelle.number_format = _WAEHRUNG
        jahr_zelle = arbeitsblatt.cell(
            zeile, _JAHRESSPALTE, f"=SUM(B{zeile}:M{zeile})"
        )
        jahr_zelle.number_format = _WAEHRUNG
        zeile += 1
    letzte_datenzeile = zeile - 1

    # Summenzeile als Formel
    summen_zeile = zeile
    arbeitsblatt.cell(zeile, 1, f"Summe {titel}").font = fett
    for monat in range(1, 13):
        spalte = get_column_letter(_ERSTE_MONATSSPALTE + monat - 1)
        if kategorien:
            formel = f"=SUM({spalte}{erste_datenzeile}:{spalte}{letzte_datenzeile})"
        else:
            formel = 0
        zelle = arbeitsblatt.cell(zeile, _ERSTE_MONATSSPALTE + monat - 1, formel)
        zelle.number_format = _WAEHRUNG
        zelle.font = fett
    jahr_zelle = arbeitsblatt.cell(
        zeile, _JAHRESSPALTE, f"=SUM(B{zeile}:M{zeile})"
    )
    jahr_zelle.number_format = _WAEHRUNG
    jahr_zelle.font = fett
    return zeile + 1, summen_zeile


def _mieter_block(
    arbeitsblatt, zeile: int, verbindung: sqlite3.Connection,
    haus_id: int, jahr: int,
) -> int:
    """Schreibt den Mieter-Block mit dem Status der Mietzahlungen."""
    fett = Font(bold=True)
    arbeitsblatt.cell(zeile, 1, "Mietzahlungen").font = fett
    zeile += 1

    mieter = mieter_laden(verbindung, haus_id)
    if not mieter:
        arbeitsblatt.cell(zeile, 1, "(keine Mieter erfasst)")
        return zeile + 1

    _kopfzeile(arbeitsblatt, zeile, "Mieter")
    arbeitsblatt.cell(zeile, _JAHRESSPALTE, "Status").font = fett
    zeile += 1
    for person in mieter:
        arbeitsblatt.cell(zeile, 1, person["name"])
        bezahlt = bezahlte_monate(verbindung, person["id"], jahr)
        for monat in range(1, 13):
            if monat in bezahlt:
                zelle = arbeitsblatt.cell(
                    zeile, _ERSTE_MONATSSPALTE + monat - 1, "x"
                )
                zelle.alignment = Alignment(horizontal="center")
        arbeitsblatt.cell(zeile, _JAHRESSPALTE, f"{len(bezahlt)}/12")
        zeile += 1
    return zeile


def jahres_export(verbindung: sqlite3.Connection, jahr: int) -> Path:
    """Erstellt die Excel-Jahresübersicht und liefert den Dateipfad."""
    summen = _monatssummen(verbindung, jahr)
    haeuser = objekte_laden(verbindung)
    einnahme_kategorien = kategorien_laden(verbindung, "einnahme")
    ausgabe_kategorien = kategorien_laden(verbindung, "ausgabe")

    arbeitsmappe = Workbook()
    blatt = arbeitsmappe.active
    blatt.title = f"Controlling {jahr}"

    blatt.column_dimensions["A"].width = 24
    for spalte in range(_ERSTE_MONATSSPALTE, _JAHRESSPALTE + 1):
        blatt.column_dimensions[get_column_letter(spalte)].width = 12

    zeile = 1
    titel = blatt.cell(zeile, 1, f"Controlling {jahr}")
    titel.font = Font(bold=True, size=14)
    zeile += 2

    for haus in haeuser:
        ueberschrift = blatt.cell(zeile, 1, haus["name"])
        ueberschrift.font = Font(bold=True, size=12)
        zeile += 1

        _kopfzeile(blatt, zeile, "Kategorie")
        zeile += 1

        zeile, einnahmen_summe = _kategorie_block(
            blatt, zeile, "Einnahmen", einnahme_kategorien, haus["id"], summen
        )
        zeile, ausgaben_summe = _kategorie_block(
            blatt, zeile, "Ausgaben", ausgabe_kategorien, haus["id"], summen
        )

        # Saldo-Zeile (Einnahmen - Ausgaben) als Formel
        blatt.cell(zeile, 1, "Saldo").font = Font(bold=True)
        for monat in range(1, 13):
            spalte = get_column_letter(_ERSTE_MONATSSPALTE + monat - 1)
            zelle = blatt.cell(
                zeile, _ERSTE_MONATSSPALTE + monat - 1,
                f"={spalte}{einnahmen_summe}-{spalte}{ausgaben_summe}",
            )
            zelle.number_format = _WAEHRUNG
            zelle.font = Font(bold=True)
        saldo_jahr = blatt.cell(
            zeile, _JAHRESSPALTE, f"=SUM(B{zeile}:M{zeile})"
        )
        saldo_jahr.number_format = _WAEHRUNG
        saldo_jahr.font = Font(bold=True)
        zeile += 2

        zeile = _mieter_block(blatt, zeile, verbindung, haus["id"], jahr)
        zeile += 2

    ziel = exporte_verzeichnis() / f"Controlling_{jahr}.xlsx"
    ziel.parent.mkdir(parents=True, exist_ok=True)
    arbeitsmappe.save(ziel)
    return ziel
