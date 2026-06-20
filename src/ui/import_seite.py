"""Import-Bereich: Kontoauszug-PDFs einlesen und Belege archivieren.

Unterstützt das Einlesen einzelner oder mehrerer Auszüge gleichzeitig
(Stapel-Import per Datei-Dialog oder Drag & Drop) sowie eine Excel-
Datenübernahme aus einer bestehenden Tabelle.

Beim PDF-Import werden die erkannten Buchungen in einer Vorschau
gezeigt: automatisch zugeordnete Zeilen grün, bestätigungspflichtige
gelb. Erst „Alle übernehmen" schreibt die Buchungen in die Datenbank.
"""

from __future__ import annotations

import sqlite3
from decimal import Decimal, InvalidOperation
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from src.db import buchungen, muster, stammdaten
from src.logic import lernsystem
from src.logic.belege import beleg_archivieren
from src.logic.pdf_import import (
    buchungszeilen_aus_text, rohtext_lesen, saldo_pruefen,
)
from src.ui.tabelle import tabelle_vorbereiten
from src.utils.eingaben import (
    ValidierungsFehler, betrag_formatieren, betrag_parsen, datum_anzeigen,
)

# Hintergrundfarben der Vorschau-Zeilen.
_FARBE_AUTO = QColor("#d7f0d7")      # grün — sicher zugeordnet
_FARBE_PRUEFEN = QColor("#fbf2c4")   # gelb — Bestätigung nötig
_FARBE_DUBLETTE = QColor("#f3cba6")  # orange — mögliche Dublette

# Anzeigetexte je Status.
_STATUS_TEXT = {
    "auto": "automatisch",
    "unsicher": "unsicher (Betrag prüfen)",
    "neu": "neu — bitte zuordnen",
}


# =========================================================================
# Hilfsfunktionen für Combo-Boxen in der Vorschau
# =========================================================================


def _haus_combo(
    verbindung: sqlite3.Connection, objekt_id: int | None
) -> QComboBox:
    """Erzeugt ein Auswahlfeld für Häuser mit Vorauswahl."""
    combo = QComboBox()
    combo.addItem("— bitte wählen —", None)
    for haus in stammdaten.objekte_laden(verbindung):
        combo.addItem(haus["name"], haus["id"])
    if objekt_id is not None:
        index = combo.findData(objekt_id)
        if index >= 0:
            combo.setCurrentIndex(index)
    return combo


def _kategorie_combo(
    verbindung: sqlite3.Connection, kategorie_id: int | None
) -> QComboBox:
    """Erzeugt ein Auswahlfeld für Kategorien mit Vorauswahl."""
    combo = QComboBox()
    combo.addItem("— bitte wählen —", None)
    for typ, bezeichnung in (("einnahme", "Einnahme"), ("ausgabe", "Ausgabe")):
        for kategorie in stammdaten.kategorien_laden(verbindung, typ):
            combo.addItem(f"{kategorie['name']} ({bezeichnung})", kategorie["id"])
    if kategorie_id is not None:
        index = combo.findData(kategorie_id)
        if index >= 0:
            combo.setCurrentIndex(index)
    return combo


# Spezial-Wert: vom Nutzer im Combo gewählter „Neuen Mieter anlegen"-Eintrag.
_NEUER_MIETER_MARKER = "__neu__"


def _mieter_combo_fuellen(
    combo: QComboBox,
    verbindung: sqlite3.Connection,
    objekt_id: int | None,
    auswahl_mieter_id: int | None = None,
) -> None:
    """Befüllt ein Mieter-Auswahlfeld passend zum gewählten Haus.

    Enthält am Ende einen Sondereintrag „+ Neuen Mieter anlegen …", der
    direkt aus der Auswahl heraus einen neuen Mieter zum aktuellen Haus
    erfasst — ohne Umweg über die Stammdaten.
    """
    bisher = auswahl_mieter_id if auswahl_mieter_id is not None \
        else combo.currentData()
    if bisher == _NEUER_MIETER_MARKER:
        bisher = None
    combo.blockSignals(True)
    combo.clear()
    combo.addItem("— kein Mieter —", None)
    if objekt_id is not None:
        for mieter in stammdaten.mieter_laden(verbindung, objekt_id):
            combo.addItem(mieter["name"], mieter["id"])
        combo.insertSeparator(combo.count())
        combo.addItem("+ Neuen Mieter anlegen …", _NEUER_MIETER_MARKER)
    if bisher is not None:
        index = combo.findData(bisher)
        if index >= 0:
            combo.setCurrentIndex(index)
    combo.blockSignals(False)


def _mieter_combo_aktivieren(
    combo: QComboBox, verbindung: sqlite3.Connection, parent: QWidget
) -> None:
    """Verbindet das Combo mit dem Quick-Add-Dialog für neue Mieter.

    Wird aufgerufen, nachdem das Combo gefüllt wurde. Sobald der Nutzer
    den Eintrag „+ Neuen Mieter anlegen …" wählt, wird ein Mieter-Dialog
    geöffnet, der ihm das Anlegen direkt aus dem Kontext heraus erlaubt.
    """
    def _gewaehlt(index: int) -> None:
        if combo.itemData(index) != _NEUER_MIETER_MARKER:
            return
        # In der Combo den Eintrag zurücksetzen, bis der Dialog geliefert hat.
        combo.blockSignals(True)
        combo.setCurrentIndex(0)
        combo.blockSignals(False)
        from src.ui.stammdaten_seite import MieterDialog
        # objekt_id aus dem zugeordneten Haus-Combo des Aufrufers ermitteln:
        # Wir suchen über die Eigenschaft das Haus-Combo. Wenn keines
        # gesetzt ist, brechen wir mit Hinweis ab.
        haus_combo: QComboBox | None = combo.property("haus_combo")
        objekt_id = haus_combo.currentData() if haus_combo else None
        if objekt_id is None:
            QMessageBox.information(
                parent, "Kein Haus gewählt",
                "Bitte zuerst ein Haus auswählen, dann den Mieter anlegen."
            )
            return
        dialog = MieterDialog(verbindung, objekt_id, parent=parent)
        if dialog.exec() == QDialog.Accepted:
            _mieter_combo_fuellen(combo, verbindung, objekt_id)
            # Neuer Mieter ist der letzte (vor dem Separator + Marker).
            for i in range(combo.count() - 1, -1, -1):
                data = combo.itemData(i)
                if isinstance(data, int):
                    combo.setCurrentIndex(i)
                    break
    combo.currentIndexChanged.connect(_gewaehlt)


# =========================================================================
# Sammel-Zuordnung
# =========================================================================


class SammelZuordnungDialog(QDialog):
    """Setzt Haus, Kategorie und/oder Mieter für mehrere markierte Zeilen."""

    def __init__(self, verbindung: sqlite3.Connection, parent=None) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self.setModal(True)
        self.setWindowTitle("Sammel-Zuordnung")

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Lege Haus, Kategorie und Mieter für alle markierten Zeilen "
            "fest. Felder, die du leer lässt, werden nicht angefasst."
        ))

        formular = QFormLayout()
        self._haus = QComboBox()
        self._haus.addItem("— nicht ändern —", "__keep__")
        for haus in stammdaten.objekte_laden(verbindung):
            self._haus.addItem(haus["name"], haus["id"])
        self._haus.currentIndexChanged.connect(self._mieter_aktualisieren)
        formular.addRow("Haus:", self._haus)

        self._kategorie = QComboBox()
        self._kategorie.addItem("— nicht ändern —", "__keep__")
        for typ, label in (("einnahme", "Einnahme"), ("ausgabe", "Ausgabe")):
            for kat in stammdaten.kategorien_laden(verbindung, typ):
                self._kategorie.addItem(f"{kat['name']} ({label})", kat["id"])
        formular.addRow("Kategorie:", self._kategorie)

        self._mieter = QComboBox()
        self._mieter.addItem("— nicht ändern —", "__keep__")
        self._mieter.addItem("— kein Mieter —", None)
        formular.addRow("Mieter:", self._mieter)

        layout.addLayout(formular)

        knoepfe = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        knoepfe.button(QDialogButtonBox.Ok).setText("Übernehmen")
        knoepfe.button(QDialogButtonBox.Cancel).setText("Abbrechen")
        knoepfe.accepted.connect(self.accept)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

    def _mieter_aktualisieren(self) -> None:
        objekt_id = self._haus.currentData()
        if objekt_id in (None, "__keep__"):
            return
        bisher = self._mieter.currentData()
        self._mieter.clear()
        self._mieter.addItem("— nicht ändern —", "__keep__")
        self._mieter.addItem("— kein Mieter —", None)
        for mieter in stammdaten.mieter_laden(self._verbindung, objekt_id):
            self._mieter.addItem(mieter["name"], mieter["id"])
        idx = self._mieter.findData(bisher)
        if idx >= 0:
            self._mieter.setCurrentIndex(idx)

    def auswahl(self) -> dict:
        """Liefert die gewählten Werte (oder ``"__keep__"`` für unverändert)."""
        return {
            "objekt_id": self._haus.currentData(),
            "kategorie_id": self._kategorie.currentData(),
            "mieter_id": self._mieter.currentData(),
        }


# =========================================================================
# Buchungs-Splitting
# =========================================================================


class SplitDialog(QDialog):
    """Teilt eine Buchung in mehrere Teilbuchungen mit eigenen Beträgen."""

    def __init__(
        self,
        verbindung: sqlite3.Connection,
        kandidat: dict,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self._original = kandidat
        self.setModal(True)
        self.setWindowTitle("Buchung splitten")
        self.resize(560, 280)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"Gesamtbetrag: {betrag_formatieren(abs(kandidat['betrag']))} € — "
            "bitte in mehrere Teile aufteilen. Die Summe muss übereinstimmen."
        ))

        self._tabelle = QTableWidget(0, 4)
        self._tabelle.setHorizontalHeaderLabels(
            ["Teilbetrag (€)", "Haus", "Kategorie", "Mieter"]
        )
        tabelle_vorbereiten(self._tabelle, sortierbar=False)
        layout.addWidget(self._tabelle)

        knopfzeile = QHBoxLayout()
        knopf_plus = QPushButton("Teil hinzufügen")
        knopf_plus.clicked.connect(self._zeile_hinzufuegen)
        knopf_minus = QPushButton("Letzten Teil entfernen")
        knopf_minus.clicked.connect(self._zeile_entfernen)
        knopfzeile.addWidget(knopf_plus)
        knopfzeile.addWidget(knopf_minus)
        knopfzeile.addStretch()
        layout.addLayout(knopfzeile)

        # Starte mit zwei Zeilen, befüllt mit der Ursprungs-Zuordnung.
        self._zeile_hinzufuegen()
        self._zeile_hinzufuegen()

        knoepfe = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        knoepfe.button(QDialogButtonBox.Ok).setText("Splitten")
        knoepfe.button(QDialogButtonBox.Cancel).setText("Abbrechen")
        knoepfe.accepted.connect(self._validieren)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

        self._teile: list[dict] | None = None

    def _zeile_hinzufuegen(self) -> None:
        zeile = self._tabelle.rowCount()
        self._tabelle.insertRow(zeile)
        feld_betrag = QLineEdit("0,00")
        self._tabelle.setCellWidget(zeile, 0, feld_betrag)
        haus = _haus_combo(self._verbindung, self._original.get("objekt_id"))
        self._tabelle.setCellWidget(zeile, 1, haus)
        kat = _kategorie_combo(
            self._verbindung, self._original.get("kategorie_id")
        )
        self._tabelle.setCellWidget(zeile, 2, kat)
        mieter = QComboBox()
        mieter.setProperty("haus_combo", haus)
        _mieter_combo_fuellen(
            mieter, self._verbindung,
            self._original.get("objekt_id"),
            self._original.get("mieter_id"),
        )
        _mieter_combo_aktivieren(mieter, self._verbindung, self)
        self._tabelle.setCellWidget(zeile, 3, mieter)
        haus.currentIndexChanged.connect(
            lambda _, h=haus, m=mieter:
            _mieter_combo_fuellen(m, self._verbindung, h.currentData())
        )

    def _zeile_entfernen(self) -> None:
        if self._tabelle.rowCount() > 2:
            self._tabelle.removeRow(self._tabelle.rowCount() - 1)

    def _validieren(self) -> None:
        teile: list[dict] = []
        summe = Decimal("0")
        for zeile in range(self._tabelle.rowCount()):
            try:
                betrag = betrag_parsen(
                    self._tabelle.cellWidget(zeile, 0).text()
                )
            except ValidierungsFehler as fehler:
                QMessageBox.warning(
                    self, "Eingabe ungültig",
                    f"Teil {zeile + 1}: {fehler}"
                )
                return
            if betrag <= 0:
                QMessageBox.warning(
                    self, "Eingabe ungültig",
                    f"Teil {zeile + 1}: Betrag muss positiv sein."
                )
                return
            objekt_id = self._tabelle.cellWidget(zeile, 1).currentData()
            kategorie_id = self._tabelle.cellWidget(zeile, 2).currentData()
            mieter_id = self._tabelle.cellWidget(zeile, 3).currentData()
            summe += betrag
            teile.append({
                "betrag_abs": betrag,
                "objekt_id": objekt_id,
                "kategorie_id": kategorie_id,
                "mieter_id": mieter_id,
            })

        ist = abs(self._original["betrag"])
        if abs(summe - ist) > Decimal("0.01"):
            QMessageBox.warning(
                self, "Summe stimmt nicht",
                f"Teilbeträge ergeben {betrag_formatieren(summe)} €, "
                f"erwartet {betrag_formatieren(ist)} €."
            )
            return
        self._teile = teile
        self.accept()

    def teile(self) -> list[dict] | None:
        return self._teile


# =========================================================================
# Vorschau-Dialog (Stapel-fähig)
# =========================================================================


class ImportVorschauDialog(QDialog):
    """Zeigt die erkannten Buchungen vor dem Übernehmen zur Kontrolle."""

    SPALTE_AKTION = 0
    SPALTE_DATUM = 1
    SPALTE_BETRAG = 2
    SPALTE_TEXT = 3
    SPALTE_HAUS = 4
    SPALTE_KATEGORIE = 5
    SPALTE_MIETER = 6
    SPALTE_BELEG = 7
    SPALTE_STATUS = 8
    SPALTEN_ZAHL = 9

    def __init__(
        self,
        verbindung: sqlite3.Connection,
        kandidaten: list[dict],
        pruefungen: list[tuple[str, dict]] | None = None,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self._kandidaten = list(kandidaten)
        self.setModal(True)
        self.setWindowTitle("Import-Vorschau")
        self.resize(1100, 560)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"{len(self._kandidaten)} Buchung(en) erkannt. "
            "Grüne Zeilen sind automatisch zugeordnet, gelbe brauchen Haus "
            "und Kategorie. Mehrere Zeilen markieren (Strg/Shift-Klick) "
            "und „Sammel-Zuordnung …“ klicken, oder das ⌖-Symbol "
            "für eine Aufteilung in mehrere Teilbuchungen."
        ))

        if pruefungen:
            for dateiname, pruefung in pruefungen:
                layout.addWidget(self._saldo_label(dateiname, pruefung))

        self._tabelle = QTableWidget(0, self.SPALTEN_ZAHL)
        self._tabelle.setHorizontalHeaderLabels([
            "", "Datum", "Betrag", "Empfänger / Zweck",
            "Haus", "Kategorie", "Mieter", "Beleg", "Status",
        ])
        tabelle_vorbereiten(self._tabelle, sortierbar=False)
        self._tabelle.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tabelle.setSelectionMode(QAbstractItemView.ExtendedSelection)
        layout.addWidget(self._tabelle)

        self._belege: dict[int, str | None] = {}
        self._tabelle_neu_aufbauen()

        knopfzeile = QHBoxLayout()
        knopf_sammel = QPushButton("Sammel-Zuordnung …")
        knopf_sammel.clicked.connect(self._sammel_zuordnen)
        knopfzeile.addWidget(knopf_sammel)
        knopfzeile.addStretch()
        layout.addLayout(knopfzeile)

        knoepfe = QDialogButtonBox()
        self._knopf_uebernehmen = knoepfe.addButton(
            "Alle übernehmen", QDialogButtonBox.AcceptRole
        )
        knoepfe.addButton("Abbrechen", QDialogButtonBox.RejectRole)
        self._knopf_uebernehmen.clicked.connect(self._uebernehmen)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

    def _saldo_label(self, dateiname: str, pruefung: dict) -> QLabel:
        """Liefert eine grüne/rote Plausibilitätsmeldung zum Endsaldo."""
        praefix = f"{Path(dateiname).name}: " if dateiname else ""
        if (pruefung["alter_saldo"] is None
                or pruefung["neuer_saldo"] is None):
            text = (f"{praefix}Plausibilitätsprüfung nicht möglich "
                    "(Anfangs-/Endsaldo nicht erkannt).")
            farbe = "#fbf2c4"
        elif pruefung["stimmt"]:
            text = (
                f"{praefix}Saldo stimmt: "
                f"{betrag_formatieren(pruefung['alter_saldo'])}"
                f" + {betrag_formatieren(pruefung['summe_buchungen'])}"
                f" = {betrag_formatieren(pruefung['neuer_saldo'])} €"
            )
            farbe = "#d7f0d7"
        else:
            differenz = (pruefung["neuer_saldo"]
                         - (pruefung["berechneter_endsaldo"] or Decimal("0")))
            text = (
                f"{praefix}Saldo stimmt NICHT — Differenz "
                f"{betrag_formatieren(differenz)} €. "
                "Bitte vor dem Übernehmen prüfen."
            )
            farbe = "#f3cba6"
        label = QLabel(text)
        label.setStyleSheet(
            f"background-color: {farbe}; padding: 6px; border-radius: 4px;"
        )
        return label

    def _tabelle_neu_aufbauen(self) -> None:
        """Erzeugt für jeden Kandidaten genau eine Tabellenzeile."""
        self._tabelle.setRowCount(0)
        for kandidat in self._kandidaten:
            self._zeile_anlegen(kandidat)

    def _zeile_anlegen(self, kandidat: dict) -> None:
        zeile = self._tabelle.rowCount()
        self._tabelle.insertRow(zeile)

        if kandidat.get("dublette"):
            farbe = _FARBE_DUBLETTE
        elif kandidat["status"] == "auto":
            farbe = _FARBE_AUTO
        else:
            farbe = _FARBE_PRUEFEN

        status_text = _STATUS_TEXT.get(kandidat["status"], kandidat["status"])
        if kandidat.get("dublette"):
            status_text += " — mögliche Dublette"

        knopf_split = QToolButton()
        knopf_split.setText("⌖")
        knopf_split.setToolTip("Buchung in mehrere Teile splitten")
        knopf_split.clicked.connect(lambda _, k=kandidat: self._zeile_splitten(k))
        self._tabelle.setCellWidget(zeile, self.SPALTE_AKTION, knopf_split)

        datum_item = QTableWidgetItem(datum_anzeigen(kandidat["datum"]))
        betrag_item = QTableWidgetItem(
            f"{betrag_formatieren(kandidat['betrag'])} €"
        )
        text_item = QTableWidgetItem(kandidat["text"])
        for spalte, item in (
            (self.SPALTE_DATUM, datum_item),
            (self.SPALTE_BETRAG, betrag_item),
            (self.SPALTE_TEXT, text_item),
        ):
            item.setBackground(farbe)
            self._tabelle.setItem(zeile, spalte, item)

        status_item = QTableWidgetItem(status_text)
        status_item.setBackground(farbe)
        self._tabelle.setItem(zeile, self.SPALTE_STATUS, status_item)

        haus_combo = _haus_combo(self._verbindung, kandidat["objekt_id"])
        self._tabelle.setCellWidget(zeile, self.SPALTE_HAUS, haus_combo)
        self._tabelle.setCellWidget(
            zeile, self.SPALTE_KATEGORIE,
            _kategorie_combo(self._verbindung, kandidat["kategorie_id"]),
        )

        mieter_combo = QComboBox()
        mieter_combo.setProperty("haus_combo", haus_combo)
        _mieter_combo_fuellen(
            mieter_combo, self._verbindung,
            kandidat["objekt_id"], kandidat.get("mieter_id"),
        )
        _mieter_combo_aktivieren(mieter_combo, self._verbindung, self)
        self._tabelle.setCellWidget(zeile, self.SPALTE_MIETER, mieter_combo)
        haus_combo.currentIndexChanged.connect(
            lambda _, hk=haus_combo, mk=mieter_combo:
            _mieter_combo_fuellen(mk, self._verbindung, hk.currentData())
        )

        # Beleg-Auswahl (Klick → Datei wählen).
        beleg_knopf = QToolButton()
        beleg_knopf.setText("📎")
        beleg_knopf.setToolTip("Beleg für diese Buchung wählen")
        beleg_knopf.clicked.connect(lambda _, k=kandidat: self._beleg_waehlen(k))
        self._tabelle.setCellWidget(zeile, self.SPALTE_BELEG, beleg_knopf)

    # --- Sammelaktionen --------------------------------------------------

    def _markierte_zeilen(self) -> list[int]:
        return sorted({idx.row() for idx in self._tabelle.selectedIndexes()})

    def _sammel_zuordnen(self) -> None:
        zeilen = self._markierte_zeilen()
        if not zeilen:
            QMessageBox.information(
                self, "Nichts markiert",
                "Bitte zuerst eine oder mehrere Zeilen markieren."
            )
            return
        dialog = SammelZuordnungDialog(self._verbindung, parent=self)
        if dialog.exec() != QDialog.Accepted:
            return
        auswahl = dialog.auswahl()
        for zeile in zeilen:
            if auswahl["objekt_id"] != "__keep__":
                haus = self._tabelle.cellWidget(zeile, self.SPALTE_HAUS)
                idx = haus.findData(auswahl["objekt_id"])
                if idx >= 0:
                    haus.setCurrentIndex(idx)
            if auswahl["kategorie_id"] != "__keep__":
                kat = self._tabelle.cellWidget(zeile, self.SPALTE_KATEGORIE)
                idx = kat.findData(auswahl["kategorie_id"])
                if idx >= 0:
                    kat.setCurrentIndex(idx)
            if auswahl["mieter_id"] != "__keep__":
                mt = self._tabelle.cellWidget(zeile, self.SPALTE_MIETER)
                idx = mt.findData(auswahl["mieter_id"])
                if idx >= 0:
                    mt.setCurrentIndex(idx)

    def _zeile_splitten(self, kandidat: dict) -> None:
        try:
            position = self._kandidaten.index(kandidat)
        except ValueError:
            return
        dialog = SplitDialog(self._verbindung, kandidat, parent=self)
        if dialog.exec() != QDialog.Accepted:
            return
        teile = dialog.teile() or []
        if not teile:
            return
        vorzeichen = -1 if kandidat["betrag"] < 0 else 1
        neue_kandidaten = []
        for index, teil in enumerate(teile, start=1):
            neue_kandidaten.append({
                "datum": kandidat["datum"],
                "betrag": teil["betrag_abs"] * vorzeichen,
                "text": f"{kandidat['text']} (Teil {index}/{len(teile)})",
                "norm": kandidat.get("norm", ""),
                "objekt_id": teil["objekt_id"],
                "kategorie_id": teil["kategorie_id"],
                "mieter_id": teil["mieter_id"],
                "status": "neu",
                "dublette": False,
            })
        self._kandidaten[position:position + 1] = neue_kandidaten
        self._tabelle_neu_aufbauen()

    def _beleg_waehlen(self, kandidat: dict) -> None:
        try:
            position = self._kandidaten.index(kandidat)
        except ValueError:
            return
        pfad, _ = QFileDialog.getOpenFileName(
            self, "Beleg auswählen", "",
            "PDF/Bild (*.pdf *.jpg *.jpeg *.png);;Alle Dateien (*)"
        )
        if pfad:
            self._belege[position] = pfad
            knopf = self._tabelle.cellWidget(position, self.SPALTE_BELEG)
            if isinstance(knopf, QToolButton):
                knopf.setText("📎 " + Path(pfad).name[:14])

    # --- Übernehmen ------------------------------------------------------

    def _uebernehmen(self) -> None:
        """Prüft die Zuordnung und schreibt alle Buchungen in die Datenbank."""
        zuordnung: list[tuple[int, int, int | None]] = []
        for zeile in range(len(self._kandidaten)):
            objekt_id = self._tabelle.cellWidget(
                zeile, self.SPALTE_HAUS
            ).currentData()
            kategorie_id = self._tabelle.cellWidget(
                zeile, self.SPALTE_KATEGORIE
            ).currentData()
            mieter_id = self._tabelle.cellWidget(
                zeile, self.SPALTE_MIETER
            ).currentData()
            if objekt_id is None or kategorie_id is None:
                QMessageBox.warning(
                    self, "Zuordnung unvollständig",
                    f"Zeile {zeile + 1}: Bitte Haus und Kategorie wählen."
                )
                return
            zuordnung.append((objekt_id, kategorie_id, mieter_id))

        try:
            for index, (kandidat, (objekt_id, kategorie_id, mieter_id)) in \
                    enumerate(zip(self._kandidaten, zuordnung)):
                beleg_pfad = None
                beleg_quelle = self._belege.get(index)
                if beleg_quelle:
                    jahr = int(kandidat["datum"][:4]) \
                        if kandidat.get("datum") else 2024
                    beleg_pfad = beleg_archivieren(beleg_quelle, jahr)
                buchungen.buchung_anlegen(
                    self._verbindung,
                    kandidat["datum"],
                    abs(kandidat["betrag"]),
                    objekt_id,
                    kategorie_id,
                    kandidat["text"],
                    beleg_pfad,
                    "import",
                    mieter_id=mieter_id,
                )
                erkennung = lernsystem.erkennungstext_bilden(
                    kandidat.get("norm", "")
                )
                if erkennung:
                    muster.muster_speichern(
                        self._verbindung, erkennung,
                        objekt_id, kategorie_id, mieter_id=mieter_id,
                    )
        except (sqlite3.Error, OSError) as fehler:
            QMessageBox.critical(self, "Fehler beim Übernehmen", str(fehler))
            return

        QMessageBox.information(
            self, "Import abgeschlossen",
            f"{len(self._kandidaten)} Buchung(en) wurden übernommen."
        )
        self.accept()


# =========================================================================
# Belege-Zuordnen-Dialog (bestehende Buchungen)
# =========================================================================


class BelegZuordnenDialog(QDialog):
    """Ordnet eine Belegdatei einer bestehenden Buchung zu."""

    def __init__(
        self,
        verbindung: sqlite3.Connection,
        datei_pfad: str,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self._datei_pfad = datei_pfad
        self.setModal(True)
        self.setWindowTitle("Beleg einer Buchung zuordnen")
        self.resize(820, 420)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"Datei: {Path(datei_pfad).name}\n"
            "Wähle die Buchung, der dieser Beleg zugeordnet werden soll."
        ))

        self._tabelle = QTableWidget(0, 5)
        self._tabelle.setHorizontalHeaderLabels(
            ["Datum", "Betrag", "Haus", "Kategorie", "Beschreibung"]
        )
        tabelle_vorbereiten(self._tabelle, sortierbar=False)
        layout.addWidget(self._tabelle)

        self._buchungen = buchungen.buchungen_laden(verbindung, beleg=False)
        self._tabelle.setRowCount(len(self._buchungen))
        for index, buchung in enumerate(self._buchungen):
            self._tabelle.setItem(
                index, 0, QTableWidgetItem(datum_anzeigen(buchung["datum"]))
            )
            self._tabelle.setItem(
                index, 1, QTableWidgetItem(
                    f"{betrag_formatieren(buchung['betrag'])} €"
                )
            )
            self._tabelle.setItem(
                index, 2, QTableWidgetItem(buchung["objekt_name"] or "—")
            )
            self._tabelle.setItem(
                index, 3, QTableWidgetItem(buchung["kategorie_name"] or "—")
            )
            self._tabelle.setItem(
                index, 4, QTableWidgetItem(buchung["beschreibung"] or "")
            )

        knoepfe = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        knoepfe.button(QDialogButtonBox.Ok).setText("Zuordnen")
        knoepfe.button(QDialogButtonBox.Cancel).setText("Abbrechen")
        knoepfe.accepted.connect(self._zuordnen)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

    def _zuordnen(self) -> None:
        zeile = self._tabelle.currentRow()
        if zeile < 0:
            QMessageBox.information(self, "Keine Buchung gewählt",
                                    "Bitte zuerst eine Buchung auswählen.")
            return
        buchung = self._buchungen[zeile]
        jahr = int(buchung["datum"][:4]) if buchung["datum"] else 2024
        try:
            relativer_pfad = beleg_archivieren(self._datei_pfad, jahr)
            buchungen.buchung_beleg_setzen(
                self._verbindung, buchung["id"], relativer_pfad
            )
        except (OSError, FileNotFoundError) as fehler:
            QMessageBox.critical(self, "Beleg-Fehler", str(fehler))
            return
        except sqlite3.Error as fehler:
            QMessageBox.critical(self, "Datenbankfehler", str(fehler))
            return
        QMessageBox.information(
            self, "Beleg archiviert",
            f"Der Beleg wurde abgelegt unter:\n{relativer_pfad}"
        )
        self.accept()


# =========================================================================
# Import-Seite (Drag & Drop, Stapel-Import, Excel-Wizard)
# =========================================================================


class ImportSeite(QWidget):
    """Einstiegsseite für PDF-Import, Excel-Import und Beleg-Archivierung."""

    def __init__(self, verbindung: sqlite3.Connection, parent=None) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Hier kannst du Volksbank-Kontoauszüge (PDF) einlesen — auch "
            "mehrere auf einmal, per Datei-Dialog oder Drag & Drop —, "
            "Daten aus einer Excel-Tabelle übernehmen oder einen Beleg "
            "archivieren."
        ))

        knopf_pdf = QPushButton("Kontoauszüge importieren (PDF, mehrere möglich)")
        knopf_pdf.clicked.connect(self._kontoauszug_importieren)
        knopf_excel = QPushButton("Excel-Datei einlesen")
        knopf_excel.clicked.connect(self._excel_importieren)
        knopf_beleg = QPushButton("Einzelnen Beleg archivieren")
        knopf_beleg.clicked.connect(self._beleg_archivieren)
        knopf_match = QPushButton("Belege-Ordner auto-matchen …")
        knopf_match.clicked.connect(self._belege_matchen)

        knopfzeile = QHBoxLayout()
        knopfzeile.addWidget(knopf_pdf)
        knopfzeile.addWidget(knopf_excel)
        knopfzeile.addWidget(knopf_beleg)
        knopfzeile.addWidget(knopf_match)
        knopfzeile.addStretch()
        layout.addLayout(knopfzeile)

        self._hinweis = QLabel(
            "Tipp: PDFs lassen sich auch direkt aus dem Datei-Explorer auf "
            "diesen Bereich ziehen. Mehrere Auszüge werden zu einer "
            "gemeinsamen Vorschau zusammengefasst."
        )
        self._hinweis.setWordWrap(True)
        self._hinweis.setStyleSheet(
            "background-color: #eaf3fb; padding: 8px; border-radius: 4px;"
        )
        layout.addWidget(self._hinweis)
        layout.addStretch()

    def aktualisieren(self) -> None:
        """Die Import-Seite hält keinen eigenen Zustand vor."""

    # --- Drag & Drop -----------------------------------------------------

    def dragEnterEvent(self, event) -> None:  # noqa: N802 - Qt-API
        if any(url.toLocalFile().lower().endswith(".pdf")
               for url in event.mimeData().urls()):
            event.acceptProposedAction()

    def dropEvent(self, event) -> None:  # noqa: N802 - Qt-API
        pfade = [
            url.toLocalFile() for url in event.mimeData().urls()
            if url.toLocalFile().lower().endswith(".pdf")
        ]
        if pfade:
            self._mehrere_kontoauszuege_oeffnen(pfade)

    # --- PDF-Import ------------------------------------------------------

    def _kontoauszug_importieren(self) -> None:
        pfade, _ = QFileDialog.getOpenFileNames(
            self, "Kontoauszüge (PDF) auswählen — mehrere möglich",
            "", "PDF-Dateien (*.pdf)"
        )
        if pfade:
            self._mehrere_kontoauszuege_oeffnen(pfade)

    def _mehrere_kontoauszuege_oeffnen(self, pfade: list[str]) -> None:
        """Liest alle übergebenen Auszüge und zeigt sie gemeinsam in der Vorschau."""
        alle_kandidaten: list[dict] = []
        pruefungen: list[tuple[str, dict]] = []
        fehler: list[str] = []

        for pfad in pfade:
            try:
                rohtext = rohtext_lesen(pfad)
                zeilen = buchungszeilen_aus_text(rohtext)
            except Exception as e:  # noqa: BLE001 - Anzeige statt Absturz
                fehler.append(f"{Path(pfad).name}: {e}")
                continue
            if not zeilen:
                fehler.append(
                    f"{Path(pfad).name}: keine Buchungen erkannt."
                )
                continue
            pruefung = saldo_pruefen(rohtext, zeilen)
            pruefungen.append((pfad, pruefung))
            for zeile in zeilen:
                alle_kandidaten.append(lernsystem.klassifizieren(
                    self._verbindung, zeile["datum"], zeile["betrag"],
                    zeile["text"],
                ))

        if fehler:
            QMessageBox.warning(
                self, "Hinweise zum PDF-Import",
                "\n".join(fehler)
            )
        if not alle_kandidaten:
            return
        ImportVorschauDialog(
            self._verbindung, alle_kandidaten,
            pruefungen=pruefungen, parent=self,
        ).exec()

    # --- Excel-Import ----------------------------------------------------

    def _excel_importieren(self) -> None:
        pfad, _ = QFileDialog.getOpenFileName(
            self, "Excel-Datei auswählen", "",
            "Excel-Dateien (*.xlsx *.xlsm);;Alle Dateien (*)"
        )
        if not pfad:
            return
        dialog = ExcelImportDialog(self._verbindung, pfad, parent=self)
        dialog.exec()

    # --- Beleg-Archivierung ---------------------------------------------

    def _beleg_archivieren(self) -> None:
        pfad, _ = QFileDialog.getOpenFileName(self, "Beleg auswählen")
        if not pfad:
            return
        BelegZuordnenDialog(self._verbindung, pfad, parent=self).exec()

    def _belege_matchen(self) -> None:
        """Wählt einen Ordner mit Rechnungen aus und schlägt Zuordnungen vor."""
        ordner = QFileDialog.getExistingDirectory(
            self, "Belege-Ordner auswählen (z. B. Rechnungen Sammelordner)"
        )
        if not ordner:
            return
        BelegeMatchDialog(self._verbindung, Path(ordner), parent=self).exec()


class BelegeMatchDialog(QDialog):
    """Zeigt Auto-Match-Vorschläge zwischen Belegdateien und Buchungen."""

    def __init__(
        self,
        verbindung: sqlite3.Connection,
        belege_ordner: Path,
        parent=None,
    ) -> None:
        super().__init__(parent)
        from src.logic import beleg_text
        self._verbindung = verbindung
        self._beleg_text = beleg_text
        self._belege_ordner = belege_ordner
        self.setModal(True)
        self.setWindowTitle("Belege auto-matchen")
        self.resize(900, 480)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"Belege-Ordner: {belege_ordner}\n\n"
            "Markiere die Zeilen, denen du zustimmst, dann unten auf "
            "Markierte übernehmen klicken. "
            "Die Belege werden archiviert und an die Buchung angehängt."
        ))

        self._paare = beleg_text.match_kandidaten(verbindung, belege_ordner)
        self._tabelle = QTableWidget(len(self._paare), 4)
        self._tabelle.setHorizontalHeaderLabels(
            ["Datei", "Betrag", "Buchung", "Treffsicherheit"]
        )
        tabelle_vorbereiten(self._tabelle, sortierbar=False)
        self._tabelle.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._tabelle.setSelectionMode(QAbstractItemView.ExtendedSelection)
        for index, paar in enumerate(self._paare):
            self._tabelle.setItem(
                index, 0, QTableWidgetItem(Path(paar["datei_pfad"]).name)
            )
            self._tabelle.setItem(
                index, 1, QTableWidgetItem(
                    f"{betrag_formatieren(paar['betrag'])} €"
                )
            )
            self._tabelle.setItem(
                index, 2, QTableWidgetItem(
                    f"#{paar['buchung_id']} — {paar['buchungstext'][:60]}"
                )
            )
            self._tabelle.setItem(
                index, 3, QTableWidgetItem(f"{paar['score']}%")
            )
        layout.addWidget(self._tabelle)

        if not self._paare:
            layout.addWidget(QLabel(
                "Keine eindeutigen Zuordnungen gefunden."
            ))

        knoepfe = QDialogButtonBox()
        self._knopf_ok = knoepfe.addButton(
            "Markierte übernehmen", QDialogButtonBox.AcceptRole
        )
        self._knopf_ok.clicked.connect(self._uebernehmen)
        knoepfe.addButton("Schließen", QDialogButtonBox.RejectRole)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

    def _uebernehmen(self) -> None:
        zeilen = sorted({idx.row() for idx in self._tabelle.selectedIndexes()})
        if not zeilen:
            QMessageBox.information(self, "Nichts markiert",
                                    "Bitte Zeilen markieren.")
            return
        erfolg = 0
        for zeile in zeilen:
            paar = self._paare[zeile]
            try:
                buchung_zeile = self._verbindung.execute(
                    "SELECT datum FROM buchungen WHERE id = ?",
                    (paar["buchung_id"],),
                ).fetchone()
                if buchung_zeile is None:
                    continue
                jahr = int(buchung_zeile["datum"][:4])
                relativ = beleg_archivieren(paar["datei_pfad"], jahr)
                buchungen.buchung_beleg_setzen(
                    self._verbindung, paar["buchung_id"], relativ
                )
                self._beleg_text.beleg_text_aktualisieren(
                    self._verbindung, relativ
                )
                erfolg += 1
            except (OSError, sqlite3.Error) as fehler:
                QMessageBox.warning(self, "Übernehmen unvollständig", str(fehler))
                continue
        QMessageBox.information(
            self, "Belege zugeordnet",
            f"{erfolg} Beleg(e) erfolgreich verknüpft."
        )
        self.accept()


# =========================================================================
# Excel-Übernahme-Wizard
# =========================================================================


class ExcelImportDialog(QDialog):
    """Liest eine Excel-Datei ein und mappt Spalten auf Buchungsfelder."""

    def __init__(
        self,
        verbindung: sqlite3.Connection,
        pfad: str,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._verbindung = verbindung
        self._pfad = pfad
        self.setModal(True)
        self.setWindowTitle("Excel-Datei einlesen")
        self.resize(640, 380)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            f"Datei: {Path(pfad).name}\n\n"
            "Lege fest, welche Spalten welchen Feldern entsprechen. "
            "Die erste Zeile mit Daten ist meist Zeile 2 (Zeile 1 = Überschrift)."
        ))

        from openpyxl import load_workbook
        try:
            self._workbook = load_workbook(pfad, data_only=True)
        except Exception as fehler:  # noqa: BLE001
            QMessageBox.critical(self, "Datei nicht lesbar", str(fehler))
            self.reject()
            return

        formular = QFormLayout()
        self._sheet = QComboBox()
        for name in self._workbook.sheetnames:
            self._sheet.addItem(name)
        self._sheet.currentIndexChanged.connect(self._spalten_neu)
        formular.addRow("Arbeitsblatt:", self._sheet)

        self._startzeile = QSpinBox()
        self._startzeile.setRange(1, 1000)
        self._startzeile.setValue(2)
        formular.addRow("Erste Datenzeile:", self._startzeile)

        self._spalte_datum = QComboBox()
        self._spalte_betrag = QComboBox()
        self._spalte_text = QComboBox()
        formular.addRow("Spalte Datum:", self._spalte_datum)
        formular.addRow("Spalte Betrag:", self._spalte_betrag)
        formular.addRow("Spalte Beschreibung:", self._spalte_text)
        layout.addLayout(formular)
        self._spalten_neu()

        knoepfe = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        knoepfe.button(QDialogButtonBox.Ok).setText("In Vorschau übernehmen")
        knoepfe.button(QDialogButtonBox.Cancel).setText("Abbrechen")
        knoepfe.accepted.connect(self._einlesen)
        knoepfe.rejected.connect(self.reject)
        layout.addWidget(knoepfe)

    def _spalten_neu(self) -> None:
        blatt = self._workbook[self._sheet.currentText()]
        spalten = list(blatt.iter_rows(min_row=1, max_row=1, values_only=True))
        kopf = list(spalten[0]) if spalten else []
        eintraege = []
        for index, wert in enumerate(kopf):
            buchstabe = chr(ord("A") + index)
            text = f"{buchstabe} — {wert}" if wert else buchstabe
            eintraege.append((text, index))
        for combo in (self._spalte_datum, self._spalte_betrag, self._spalte_text):
            combo.clear()
            for text, idx in eintraege:
                combo.addItem(text, idx)

    def _einlesen(self) -> None:
        blatt = self._workbook[self._sheet.currentText()]
        start = self._startzeile.value()
        sp_datum = self._spalte_datum.currentData()
        sp_betrag = self._spalte_betrag.currentData()
        sp_text = self._spalte_text.currentData()
        if None in (sp_datum, sp_betrag, sp_text):
            QMessageBox.warning(self, "Spalten unvollständig",
                                "Bitte alle drei Spalten zuordnen.")
            return

        kandidaten: list[dict] = []
        for row in blatt.iter_rows(min_row=start, values_only=True):
            if all(z is None or z == "" for z in row):
                continue
            try:
                datum_roh = row[sp_datum]
                if hasattr(datum_roh, "strftime"):
                    datum = datum_roh.strftime("%Y-%m-%d")
                else:
                    datum = str(datum_roh)[:10]
                if not datum or len(datum) < 8:
                    continue
                betrag_roh = row[sp_betrag]
                if isinstance(betrag_roh, (int, float)):
                    betrag = Decimal(str(betrag_roh))
                else:
                    betrag = betrag_parsen(str(betrag_roh))
                text = str(row[sp_text] or "")
            except (ValueError, InvalidOperation, ValidierungsFehler):
                continue
            kandidaten.append(lernsystem.klassifizieren(
                self._verbindung, datum, betrag, text,
            ))

        if not kandidaten:
            QMessageBox.information(
                self, "Keine Datensätze",
                "In der Datei wurden keine gültigen Buchungen erkannt."
            )
            return
        self.accept()
        ImportVorschauDialog(
            self._verbindung, kandidaten, parent=self.parent(),
        ).exec()
